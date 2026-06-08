import pandas as pd
from openpyxl import load_workbook
import os
from pathlib import Path

class ExportarModel:

    def exportar_excel(self, resultados, nombre_archivo):

        resultados_ordenados = sorted(
            resultados,
            key=lambda x: str(x.get("alumno", "")).lower()
        )

        datos_excel = []

        for i, r in enumerate(resultados_ordenados, start=1):

            fila = {
                "No.": i,
                **r,
                "Firma": ""
            }

            datos_excel.append(fila)

        df = pd.DataFrame(datos_excel)

        carpeta_descargas = Path.home() / "Downloads"

        ruta_archivo = carpeta_descargas / nombre_archivo
        
        df.to_excel(
            ruta_archivo,
            index=False
        )

        wb = load_workbook(ruta_archivo)
        ws = wb.active

        for columna in ws.columns:
            longitud = max(
                len(str(celda.value))
                if celda.value is not None else 0
                for celda in columna
            )

            ws.column_dimensions[
                columna[0].column_letter
            ].width = longitud + 5

        for celda in ws[1]:
            if celda.value == "Firma":
                ws.column_dimensions[
                    celda.column_letter
                ].width = 30
                break
            
        wb.save(ruta_archivo)
        return True